#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = ROOT / "data" / "clean"
OUTPUT_XLSX = ROOT / "excel" / "dallas_crime_report.xlsx"


def autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 28)


def write_dataframe(sheet, frame: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    for col_offset, column in enumerate(frame.columns, start=start_col):
        cell = sheet.cell(row=start_row, column=col_offset, value=column)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")

    for row_offset, row in enumerate(frame.itertuples(index=False), start=start_row + 1):
        for col_offset, value in enumerate(row, start=start_col):
            sheet.cell(row=row_offset, column=col_offset, value=value)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    autosize_columns(sheet)


def main() -> int:
    division_stats_path = INPUT_DIR / "division_stats.csv"
    division_trends_path = INPUT_DIR / "division_trends.csv"
    offense_counts_path = INPUT_DIR / "offense_year_counts.csv"
    zip_demographics_path = INPUT_DIR / "zip_crime_demographics.csv"

    for path in [division_stats_path, division_trends_path, offense_counts_path]:
        if not path.exists():
            raise FileNotFoundError(f"Missing {path}. Run scripts/process_historical.py first.")

    division_stats = pd.read_csv(division_stats_path)
    division_trends = pd.read_csv(division_trends_path)
    offense_year_counts = pd.read_csv(offense_counts_path)
    zip_demographics = pd.read_csv(zip_demographics_path) if zip_demographics_path.exists() else pd.DataFrame()

    latest_year = int(division_stats["year"].max())
    baseline_year = int(division_trends["baseline_year"].iloc[0])

    workbook = Workbook()

    summary = workbook.active
    summary.title = "Monthly Crime Summary"
    summary_pivot = (
        division_stats.pivot_table(index="division", columns="year", values="total_crimes", aggfunc="sum", fill_value=0)
        .reset_index()
    )
    write_dataframe(summary, summary_pivot)
    if summary.max_column > 1 and summary.max_row > 1:
        summary.conditional_formatting.add(
            f"B2:{get_column_letter(summary.max_column)}{summary.max_row}",
            ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"),
        )

    breakdown = workbook.create_sheet("Crime Type Breakdown")
    latest_offenses = (
        offense_year_counts.loc[offense_year_counts["year"] == latest_year]
        .sort_values("incident_count", ascending=False)
        .head(40)["offense_type"]
        .tolist()
    )
    offense_pivot = (
        offense_year_counts.loc[offense_year_counts["offense_type"].isin(latest_offenses)]
        .pivot_table(index=["crime_category", "offense_type"], columns="year", values="incident_count", aggfunc="sum", fill_value=0)
        .reset_index()
    )
    latest_total = float(offense_pivot[latest_year].sum()) if latest_year in offense_pivot.columns else 0.0
    if latest_total > 0 and latest_year in offense_pivot.columns:
        offense_pivot["pct_of_latest_year"] = (offense_pivot[latest_year] / latest_total).round(4)
    write_dataframe(breakdown, offense_pivot)

    report = workbook.create_sheet("City Council Report")
    report["A1"] = "Dallas Crime Intelligence Report 2024"
    report["A1"].font = Font(size=18, bold=True)
    report["A3"] = "Executive Summary"
    report["A3"].font = Font(size=13, bold=True)

    improving = division_trends.loc[division_trends["trend_label"] == "Improving"].sort_values("crime_change_pct").head(5)
    worsening = division_trends.loc[division_trends["trend_label"] == "Worsening"].sort_values("crime_change_pct", ascending=False).head(5)
    latest_risk = division_stats.loc[division_stats["year"] == latest_year].sort_values("risk_score", ascending=False)

    key_findings = [
        f"Analyzed {int(division_stats['total_crimes'].sum()):,} incidents across {division_stats['year'].nunique()} years of cleaned history.",
        f"{len(improving):,} divisions are improving and {len(worsening):,} are worsening versus the comparison period.",
        f"Highest current risk score division: {latest_risk.iloc[0]['division']} ({latest_risk.iloc[0]['risk_score']:.1f}).",
        f"Baseline comparison spans {baseline_year} to {latest_year} for long-run trend analysis.",
    ]
    for index, finding in enumerate(key_findings, start=4):
        report[f"A{index}"] = f"- {finding}"

    report["A10"] = "Top 5 Improving Divisions"
    report["A10"].font = Font(bold=True)
    write_dataframe(report, improving[["division", "crime_change_pct", "risk_score"]], start_row=11, start_col=1)
    for row in range(12, 12 + len(improving)):
        for col in range(1, 4):
            report.cell(row=row, column=col).fill = PatternFill("solid", fgColor="E2F0D9")

    report["F10"] = "Top 5 Worsening Divisions"
    report["F10"].font = Font(bold=True)
    write_dataframe(report, worsening[["division", "crime_change_pct", "risk_score"]], start_row=11, start_col=6)
    for row in range(12, 12 + len(worsening)):
        for col in range(6, 9):
            report.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FCE4D6")

    chart = BarChart()
    chart.title = f"Risk Scores ({latest_year})"
    chart.y_axis.title = "Risk Score"
    chart.x_axis.title = "Division"
    data_ref = Reference(report, min_col=9, min_row=11, max_row=11 + len(latest_risk))
    cats_ref = Reference(report, min_col=10, min_row=12, max_row=11 + len(latest_risk))
    report["I11"] = "risk_score"
    report["J11"] = "division"
    for idx, row in enumerate(latest_risk[["risk_score", "division"]].itertuples(index=False), start=12):
        report[f"I{idx}"] = float(row.risk_score)
        report[f"J{idx}"] = row.division
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 8
    chart.width = 14
    report.add_chart(chart, "A19")
    autosize_columns(report)

    trend_sheet = workbook.create_sheet("Trend Analysis")
    trend_output = division_trends[
        [
            "division",
            f"total_crimes_{baseline_year}",
            f"total_crimes_{latest_year}",
            "pct_change_since_baseline",
            "trend_arrow",
            "risk_score",
            "trend_label",
        ]
    ].copy()
    write_dataframe(trend_sheet, trend_output)
    for row in range(2, trend_sheet.max_row + 1):
        label = trend_sheet.cell(row=row, column=trend_sheet.max_column).value
        fill = None
        if label == "Improving":
            fill = PatternFill("solid", fgColor="E2F0D9")
        elif label == "Worsening":
            fill = PatternFill("solid", fgColor="FCE4D6")
        if fill:
            for col in range(1, trend_sheet.max_column + 1):
                trend_sheet.cell(row=row, column=col).fill = fill

    if not zip_demographics.empty:
        zip_sheet = workbook.create_sheet("ZIP Demographics")
        latest_zip = (
            zip_demographics.loc[zip_demographics["year"] == zip_demographics["year"].max(), [
                "zip_code",
                "combined_risk_score",
                "incidents_per_1000",
                "poverty_rate_pct",
                "unemployment_rate_pct",
                "median_household_income",
                "bachelors_plus_rate_pct",
            ]]
            .sort_values("combined_risk_score", ascending=False)
            .head(50)
            .copy()
        )
        for column in [
            "combined_risk_score",
            "incidents_per_1000",
            "poverty_rate_pct",
            "unemployment_rate_pct",
            "median_household_income",
            "bachelors_plus_rate_pct",
        ]:
            latest_zip[column] = latest_zip[column].round(2)
        write_dataframe(zip_sheet, latest_zip)
        if zip_sheet.max_column > 1 and zip_sheet.max_row > 1:
            zip_sheet.conditional_formatting.add(
                "B2:B51",
                ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"),
            )

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        autosize_columns(sheet)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_XLSX)
    print(f"Saved {OUTPUT_XLSX.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
